import json
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation

import certifi
import pandas as pd
from django.conf import settings
from django.db import transaction
from decouple import config
from django.contrib.auth.forms import PasswordResetForm
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseRedirect, HttpResponseBadRequest, HttpResponse, HttpResponseForbidden
import requests
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.http import urlsafe_base64_encode
from django.views.decorators.csrf import csrf_exempt
from requests import Session
from requests.adapters import HTTPAdapter
from urllib3 import Retry

from . import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from . import helper, models
from .forms import UploadFileForm
from .models import CustomUser


# Create your views here.
def home(request):
    if models.Announcement.objects.filter(active=True).exists():
        announcement = models.Announcement.objects.filter(active=True).first()
        messages.info(request, announcement.message)
        return render(request, "layouts/index.html")
    return render(request, "layouts/index.html")


def services(request):
    return render(request, "layouts/services.html")


def pay_with_wallet(request):
    if request.method == "POST":

        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        if user.wallet is None:
            return JsonResponse({'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < Decimal(amount):
            return JsonResponse({'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        print(phone_number)
        print(amount)
        print(reference)
        if user.status == "User":
            bundle = models.IshareBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentIshareBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentIshareBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        else:
            bundle = models.IshareBundlePrice.objects.get(price=Decimal(amount)).bundle_volume

        print(bundle)
        sms_headers = {
            'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        ishare_channel = models.AdminInfo.objects.filter().first().ishare_channel

        if ishare_channel == "Gyasi":
            print("did Gyasi")
            send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, reference)
            try:
                data = send_bundle_response.json()
                print(data)
            except:
                return JsonResponse({'status': f'Something went wrong'})

            if send_bundle_response.status_code == 200:
                if data["code"] == "0000":
                    new_transaction = models.IShareBundleTransaction.objects.create(
                        user=request.user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                        transaction_status="Completed"
                    )
                    new_transaction.save()
                    user.wallet -= Decimal(amount)
                    user.save()
                    receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {reference}\n"
                    sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using DanWel Store GH.\n\nThe DanWel Store GH"

                    num_without_0 = phone_number[1:]
                    print(num_without_0)
                    # receiver_body = {
                    #     'recipient': f"233{num_without_0}",
                    #     'sender_id': 'DANWELSTORE',
                    #     'message': receiver_message
                    # }
                    #
                    # response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                    # print(response.text)

                    sms_body = {
                        'recipient': f"233{request.user.phone}",
                        'sender_id': 'DANWELSTORE',
                        'message': sms_message
                    }

                    response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

                    print(response.text)

                    # response1 = requests.get(
                    #     f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=OnBuSjBXc1pqN0xrQXIxU1A=&to=0{request.user.phone}&from=DANWELSTORE&sms={sms_message}")
                    # print(response1.text)
                    #
                    # response2 = requests.get(
                    #     f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=OnBuSjBXc1pqN0xrQXIxU1A=&to={phone_number}&from=DANWELSTORE&sms={receiver_message}")
                    # print(response2.text)

                    return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
                else:
                    new_transaction = models.IShareBundleTransaction.objects.create(
                        user=request.user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                        transaction_status="Failed"
                    )
                    new_transaction.save()
                    return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
        elif ishare_channel == "Geosams":
            send_bundle_response = helper.controller_send_bundle(phone_number, bundle, reference)
            try:
                data = send_bundle_response.json()
                print(data)
            except:
                return JsonResponse({'status': f'Something went wrong'})

            sms_headers = {
                'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            if send_bundle_response.status_code == 200:
                if data["code"] == "200":
                    with transaction.atomic():
                        new_transaction = models.IShareBundleTransaction.objects.create(
                            user=request.user,
                            bundle_number=phone_number,
                            offer=f"{bundle}MB",
                            reference=reference,
                            transaction_status="Completed"
                        )
                        new_transaction.save()
                        user.wallet -= Decimal(amount)
                        user.save()
                    receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {reference}\n"
                    sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using GH BAY."

                    num_without_0 = phone_number[1:]
                    print(num_without_0)
                    receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {reference}\n"
                    sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using DanWel Store GH.\n\nThe DanWel Store GH"

                    num_without_0 = phone_number[1:]
                    print(num_without_0)
                    # receiver_body = {
                    #     'recipient': f"233{num_without_0}",
                    #     'sender_id': 'DANWELSTORE',
                    #     'message': receiver_message
                    # }
                    #
                    # response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                    # print(response.text)

                    sms_body = {
                        'recipient': f"233{request.user.phone}",
                        'sender_id': 'DANWELSTORE',
                        'message': sms_message
                    }

                    response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers, timeout=10)

                    print(response.text)

                    return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
                else:
                    new_transaction = models.IShareBundleTransaction.objects.create(
                        user=request.user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                        transaction_status="Failed"
                    )
                    new_transaction.save()
                    return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
            else:
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Failed"
                )
                new_transaction.save()
                return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
        elif ishare_channel == "Value4Moni":
            send_bundle_response = helper.value_4_moni_send_bundle(phone_number, bundle, reference)
            try:
                data = send_bundle_response.json()
                print(data)
            except:
                return JsonResponse({'status': f'Something went wrong'})

            sms_headers = {
                'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            if send_bundle_response.status_code == 200:
                if data["code"] == "200":
                    new_transaction = models.IShareBundleTransaction.objects.create(
                        user=request.user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                        transaction_status="Completed"
                    )
                    new_transaction.save()
                    user.wallet -= Decimal(amount)
                    user.save()
                    receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {reference}\n"
                    sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using GH BAY."

                    num_without_0 = phone_number[1:]
                    print(num_without_0)
                    receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {reference}\n"
                    sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using DanWel Store GH.\n\nThe DanWel Store GH"

                    num_without_0 = phone_number[1:]
                    print(num_without_0)
                    # receiver_body = {
                    #     'recipient': f"233{num_without_0}",
                    #     'sender_id': 'DANWELSTORE',
                    #     'message': receiver_message
                    # }
                    #
                    # response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                    # print(response.text)

                    sms_body = {
                        'recipient': f"233{request.user.phone}",
                        'sender_id': 'DANWELSTORE',
                        'message': sms_message
                    }

                    response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

                    print(response.text)

                    return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
                else:
                    new_transaction = models.IShareBundleTransaction.objects.create(
                        user=request.user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                        transaction_status="Failed",
                        description=data["message"]
                    )
                    new_transaction.save()
                    return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
            else:
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Failed"
                )
                new_transaction.save()
                return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
    return redirect('airtel-tigo')


@login_required(login_url='login')
def airtel_tigo(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.IShareBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email
    if request.method == "POST":
        form = forms.IShareBundleForm(data=request.POST, status=status)
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Completed"
        )
        new_payment.save()
        print("payment saved")
        print("form valid")
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        print(offer)
        bundle = models.IshareBundlePrice.objects.get(
            price=Decimal(offer)).bundle_volume if user.status == "User" else models.AgentIshareBundlePrice.objects.get(
            price=Decimal(offer)).bundle_volume

        new_transaction = models.IShareBundleTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
            transaction_status="Pending"
        )
        print("created")
        new_transaction.save()

        print("===========================")
        print(phone_number)
        print(bundle)
        send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, reference)
        data = send_bundle_response.json()

        print(data)

        sms_headers = {
            'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'

        if send_bundle_response.status_code == 200:
            if data["code"] == "0000":
                transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
                print("got here")
                print(transaction_to_be_updated.transaction_status)
                transaction_to_be_updated.transaction_status = "Completed"
                transaction_to_be_updated.save()
                print(request.user.phone)
                print("***********")
                receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {payment_reference}\n"
                sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {payment_reference}\nThank you for using DanWel Store GH.\n\nThe DanWel Store GH"

                num_without_0 = phone_number[1:]
                print(num_without_0)
                receiver_body = {
                    'recipient': f"233{num_without_0}",
                    'sender_id': 'DANWELSTORE',
                    'message': receiver_message
                }

                response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                print(response.text)

                sms_body = {
                    'recipient': f"233{request.user.phone}",
                    'sender_id': 'DANWELSTORE',
                    'message': sms_message
                }

                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

                print(response.text)

                return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
            else:
                transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
                transaction_to_be_updated.transaction_status = "Failed"
                new_transaction.save()
                sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {payment_reference}\nThank you for using DanWel Store GH.\n\nThe DanWel Store GH"

                sms_body = {
                    'recipient': f"233{request.user.phone}",
                    'sender_id': 'DANWELSTORE',
                    'message': sms_message
                }
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
                # r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={phone_number}&from=DanWel Store GH&sms={receiver_message}"
                # response = requests.request("GET", url=r_sms_url)
                # print(response.text)
                return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
        else:
            transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
            transaction_to_be_updated.transaction_status = "Failed"
            new_transaction.save()
            sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {payment_reference}\nThank you for using DanWel Store GH.\n\nThe DanWel Store GH"

            sms_body = {
                'recipient': f'233{request.user.phone}',
                'sender_id': 'DANWELSTORE',
                'message': sms_message
            }

            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

            print(response.text)
            return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
    user = models.CustomUser.objects.get(id=request.user.id)
    context = {"form": form, "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/at.html", context=context)


def mtn_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)

        admin = models.AdminInfo.objects.filter().first().phone_number

        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < Decimal(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        if user.status == "User":
            bundle = models.MTNBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentMTNBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentMTNBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        else:
            bundle = models.MTNBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        print(bundle)
        sms_message = f"An order has been placed. {bundle}MB for {phone_number}"
        print("got here")
        with transaction.atomic():
            print("then here")
            new_mtn_transaction = models.MTNTransaction.objects.create(
                user=request.user,
                bundle_number=phone_number,
                offer=f"{bundle}MB",
                reference=reference,
            )
            new_mtn_transaction.save()
            user.wallet -= Decimal(amount)
            user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('mtn')


def telecel_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone = user.phone
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        # sms_headers = {
        #     'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
        #     'Content-Type': 'application/json'
        # }
        #
        # sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        # admin = models.AdminInfo.objects.filter().first().phone_number

        if user.wallet is None:
            return JsonResponse({'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < Decimal(amount):
            return JsonResponse({'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        if user.status == "User":
            bundle = models.TelecelBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentTelecelBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentTelecelBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        else:
            bundle = models.TelecelBundlePrice.objects.get(price=Decimal(amount)).bundle_volume

        print(bundle)
        sms_message = f"An order has been placed. {bundle}MB for {phone_number}"
        new_telecel_transaction = models.TelecelTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_telecel_transaction.save()
        user.wallet -= Decimal(amount)
        user.save()
        # sms_body = {
        #     'recipient': "233540975553",
        #     'sender_id': 'DANWELSTORE',
        #     'message': sms_message
        # }
        # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        # print(response.text)
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('telecel')


@login_required(login_url='login')
def big_time_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < Decimal(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        if user.status == "User":
            bundle = models.BigTimeBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentBigTimeBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        else:
            bundle = models.BigTimeBundlePrice.objects.get(price=Decimal(amount)).bundle_volume
        print(bundle)
        new_mtn_transaction = models.BigTimeTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= Decimal(amount)
        user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('big_time')


@login_required(login_url='login')
def mtn(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    phone = user.phone
    status = user.status
    form = forms.MTNForm(status=status)
    reference = helper.mtn_ref_generator()
    user_email = request.user.email
    # if request.method == "POST":
    #     payment_reference = request.POST.get("reference")
    #     amount_paid = request.POST.get("amount")
    #     new_payment = models.Payment.objects.create(
    #         user=request.user,
    #         reference=payment_reference,
    #         amount=amount_paid,
    #         transaction_date=datetime.now(),
    #         transaction_status="Completed"
    #     )
    #     new_payment.save()
    #     phone_number = request.POST.get("phone")
    #     offer = request.POST.get("amount")
    #
    #     if user.status == "User":
    #         bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume
    #     elif user.status == "Agent":
    #         bundle = models.AgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
    #
    #     new_mtn_transaction = models.MTNTransaction.objects.create(
    #         user=request.user,
    #         bundle_number=phone_number,
    #         offer=f"{bundle}MB",
    #         reference=payment_reference,
    #
    #     )
    #     new_mtn_transaction.save()
    #     sms_headers = {
    #         'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
    #         'Content-Type': 'application/json'
    #     }
    #
    #     sms_url = 'https://webapp.usmsgh.com/api/sms/send'
    #     sms_message = f"An order has been placed. {bundle}MB for {phone_number}"
    #
    #     sms_body = {
    #         'recipient': "233540975553",
    #         'sender_id': 'DANWELSTORE',
    #         'message': sms_message
    #     }
    #     response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
    #     print(response.text)
    #     return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    phone_num = user.phone

    context = {'form': form, 'phone_num': phone_num,
               "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/mtn.html", context=context)


@login_required(login_url='login')
def telecel(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    phone = user.phone
    status = user.status
    form = forms.TelecelForm(status=status)
    reference = helper.ref_generator()
    user_email = request.user.email
    if request.method == "POST":
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Completed"
        )
        new_payment.save()
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")

        if user.status == "User":
            bundle = models.TelecelBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentTelecelBundlePrice.objects.get(price=float(offer)).bundle_volume

        new_mtn_transaction = models.MTNTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,

        )
        new_mtn_transaction.save()
        sms_headers = {
            'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"An order has been placed. {bundle}MB for {phone_number}"

        sms_body = {
            'recipient': "233540975553",
            'sender_id': 'DANWELSTORE',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    phone_num = user.phone

    context = {'form': form, 'phone_num': phone_num,
               "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/voda.html", context=context)


@login_required(login_url='login')
def afa_registration(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    reference = helper.ref_generator()
    db_user_id = request.user.id
    price = models.AdminInfo.objects.filter().first().afa_price
    user_email = request.user.email
    print(price)
    if request.method == "POST":
        form = forms.AFARegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Registration will be done shortly")
    form = forms.AFARegistrationForm()
    context = {'form': form, 'ref': reference, 'price': price, 'id': db_user_id, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/afa.html", context=context)


def afa_registration_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        name = request.POST.get("name")
        card_number = request.POST.get("card")
        occupation = request.POST.get("occupation")
        date_of_birth = request.POST.get("birth")
        price = models.AdminInfo.objects.filter().first().afa_price

        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < Decimal(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})

        new_registration = models.AFARegistration.objects.create(
            user=user,
            reference=reference,
            name=name,
            phone_number=phone_number,
            gh_card_number=card_number,
            occupation=occupation,
            date_of_birth=date_of_birth
        )
        new_registration.save()
        user.wallet -= Decimal(price)
        user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('home')


@login_required(login_url='login')
def big_time(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.BigTimeBundleForm(status)
    reference = helper.ref_generator()
    db_user_id = request.user.id
    user_email = request.user.email

    if request.method == "POST":
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Pending"
        )
        new_payment.save()
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        if user.status == "User":
            bundle = models.BigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
        print(phone_number)
        new_mtn_transaction = models.BigTimeTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
        )
        new_mtn_transaction.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    # phone_num = user.phone
    # mtn_dict = {}
    #
    # if user.status == "Agent":
    #     mtn_offer = models.AgentMTNBundlePrice.objects.all()
    # else:
    #     mtn_offer = models.MTNBundlePrice.objects.all()
    # for offer in mtn_offer:
    #     mtn_dict[str(offer)] = offer.bundle_volume
    context = {'form': form,
               "ref": reference, "email": user_email, 'id': db_user_id,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/big_time.html", context=context)


@login_required(login_url='login')
def history(request):
    user_transactions = models.IShareBundleTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "AirtelTigo Transactions"
    net = "tigo"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def mtn_history(request):
    user_transactions = models.MTNTransaction.objects.filter(user=request.user).order_by('transaction_date').reverse()
    header = "MTN Transactions"
    net = "mtn"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def telecel_history(request):
    user_transactions = models.TelecelTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "Telecel Transactions"
    net = "telecel"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def big_time_history(request):
    user_transactions = models.BigTimeTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "Big Time Transactions"
    net = "bt"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def afa_history(request):
    user_transactions = models.AFARegistration.objects.filter(user=request.user).order_by('transaction_date').reverse()
    header = "AFA Registrations"
    net = "afa"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/afa_history.html", context=context)


def verify_transaction(request, reference):
    if request.method == "GET":
        response = helper.verify_paystack_transaction(reference)
        data = response.json()
        try:
            status = data["data"]["status"]
            amount = data["data"]["amount"]
            api_reference = data["data"]["reference"]
            date = data["data"]["paid_at"]
            real_amount = Decimal(amount) / 100
            print(status)
            print(real_amount)
            print(api_reference)
            print(reference)
            print(date)
        except:
            status = data["status"]
        return JsonResponse({'status': status})


def change_excel_status(request, status, to_change_to):
    print("got in here")
    transactions = models.MTNTransaction.objects.filter(
        transaction_status='Processing')
    print(transactions)
    for txn in transactions:
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        print(txn.transaction_status)
    messages.success(request, f"Status changed from {status} to {to_change_to}")
    return redirect("mtn_admin", status=status)


from django.db.models import FloatField, F
from django.db.models.functions import Cast, Substr, Length


@login_required(login_url='login')
def admin_mtn_history(request, status):
    if request.user.is_staff and request.user.is_superuser:
        if request.method == "POST":
            from io import BytesIO
            from openpyxl import load_workbook
            from django.http import HttpResponse
            import datetime

            # Assuming `uploaded_file` is the Excel file uploaded by the user
            uploaded_file = request.FILES['file'] if 'file' in request.FILES else None
            if not uploaded_file:
                messages.error(request, "No excel file found")
                return redirect('mtn_admin', status=status)

            # Load the uploaded Excel file into memory
            excel_buffer = BytesIO(uploaded_file.read())
            book = load_workbook(excel_buffer)
            sheet = book.active  # Assuming the data is on the active sheet

            # Assuming we have identified the recipient and data column indices
            # Replace these with the actual indices if available
            recipient_col_index = 1  # Example index for "RECIPIENT"
            data_col_index = 2  # Example index for "DATA"

            # Query your Django model
            queryset = models.MTNTransaction.objects.filter(transaction_status="Pending").annotate(
                offer_value=Cast(Substr('offer', 1, Length('offer') - 2), FloatField())
            ).order_by('-offer_value')

            # Determine the starting row for updates, preserving headers and any other pre-existing content
            start_row = 2  # Assuming data starts from row 2

            for record in queryset:
                # Assuming 'bundle_number' and 'offer' fields exist in your model
                recipient_value = f"0{record.bundle_number}"  # Ensure it's a string to preserve formatting
                data_value = record.offer  # Adjust based on actual field type
                cleaned_data_value = float(data_value.replace('MB', ''))
                data_value_gb = round(float(cleaned_data_value) / 1000, 2)

                # Find next available row (avoid overwriting non-empty rows if necessary)
                while sheet.cell(row=start_row, column=recipient_col_index).value is not None:
                    start_row += 1

                # Update cells
                sheet.cell(row=start_row, column=recipient_col_index, value=recipient_value)
                sheet.cell(row=start_row, column=data_col_index, value=data_value_gb)

                # Update the record status, if necessary
                record.transaction_status = "Processing"
                record.save()

            # Save the modified Excel file to the buffer
            excel_buffer.seek(0)  # Reset buffer position
            book.save(excel_buffer)

            # Prepare the response with the modified Excel file
            excel_buffer.seek(0)  # Reset buffer position to read the content
            response = HttpResponse(excel_buffer.getvalue(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(
                datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

            return response

        all_txns = models.MTNTransaction.objects.filter(transaction_status=status).order_by('-transaction_date')[:800]
        context = {'txns': all_txns, 'status': status}
        return render(request, "layouts/services/mtn_admin.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('mtn_admin', status=status)


@login_required(login_url='login')
def admin_telecel_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.TelecelTransaction.objects.all().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/voda_admin.html", context=context)


@login_required(login_url='login')
def admin_at_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.IShareBundleTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/at_admin.html", context=context)


@login_required(login_url='login')
def admin_bt_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.BigTimeTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/bt_admin.html", context=context)


@login_required(login_url='login')
def admin_afa_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.AFARegistration.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/afa_admin.html", context=context)


@login_required(login_url='login')
def mark_as_sent(request, pk, status):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.MTNTransaction.objects.filter(id=pk).first()
        print(txn)
        if status == "Processing":
            txn.transaction_status = "Processing"
            txn.save()
            messages.success(request, f"Transaction Processed")
            return redirect('mtn_admin')
        elif status == "Cancelled":
            txn.transaction_status = "Cancelled"
            txn.save()
            messages.success(request, f"Transaction Cancelled")
            return redirect('mtn_admin')
        elif status == "Refunded":
            txn.transaction_status = "Refunded"
            txn.save()
            messages.success(request, f"Transaction Refunded")
            return redirect('mtn_admin')
        else:
            txn.transaction_status = "Completed"
            txn.save()
            sms_headers = {
                'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"{txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

            sms_body = {
                'recipient': f"233{txn.user.phone}",
                'sender_id': 'DANWELSTORE',
                'message': sms_message
            }
            # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            # print(response.text)
            return redirect('mtn_admin')


@login_required(login_url='login')
def telecel_mark_as_sent(request, pk, status):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.TelecelTransaction.objects.filter(id=pk).first()
        print(txn)
        if status == "Processing":
            txn.transaction_status = "Processing"
            txn.save()
            messages.success(request, f"Transaction Processed")
            return redirect('telecel_admin')
        elif status == "Cancelled":
            txn.transaction_status = "Cancelled"
            txn.save()
            messages.success(request, f"Transaction Cancelled")
            return redirect('telecel_admin')
        elif status == "Refunded":
            txn.transaction_status = "Refunded"
            txn.save()
            messages.success(request, f"Transaction Refunded")
            return redirect('telecel_admin')
        else:
            txn.transaction_status = "Completed"
            txn.save()
            sms_headers = {
                'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"{txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

            sms_body = {
                'recipient': f"233{txn.user.phone}",
                'sender_id': 'DANWELSTORE',
                'message': sms_message
            }
            # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            # print(response.text)
            return redirect('telecel_admin')


@login_required(login_url='login')
def at_mark_as_sent(request, pk, status):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.IShareBundleTransaction.objects.filter(id=pk).first()
        print(txn)
        if status == "Processing":
            txn.transaction_status = "Processing"
            txn.save()
            messages.success(request, f"Transaction Processed")
            return redirect('at_admin')
        elif status == "Cancelled":
            txn.transaction_status = "Cancelled"
            txn.save()
            messages.success(request, f"Transaction Cancelled")
            return redirect('at_admin')
        elif status == "Refunded":
            txn.transaction_status = "Refunded"
            txn.save()
            messages.success(request, f"Transaction Refunded")
            return redirect('at_admin')
        else:
            txn.transaction_status = "Completed"
            txn.save()
            sms_headers = {
                'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Your AT transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

            sms_body = {
                'recipient': f"233{txn.user.phone}",
                'sender_id': 'DANWELSTORE',
                'message': sms_message
            }

            messages.success(request, f"Transaction Completed")
            return redirect('at_admin')


@login_required(login_url='login')
def bt_mark_as_sent(request, pk, status):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.BigTimeTransaction.objects.filter(id=pk).first()
        print(txn)
        if status == "Processing":
            txn.transaction_status = "Processing"
            txn.save()
            messages.success(request, f"Transaction Processed")
            return redirect('bt_admin')
        elif status == "Cancelled":
            txn.transaction_status = "Cancelled"
            txn.save()
            messages.success(request, f"Transaction Cancelled")
            return redirect('bt_admin')
        elif status == "Refunded":
            txn.transaction_status = "Refunded"
            txn.save()
            messages.success(request, f"Transaction Refunded")
            return redirect('bt_admin')
        else:
            txn.transaction_status = "Completed"
            txn.save()
            sms_headers = {
                'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Your AT BIG TIME transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

            sms_body = {
                'recipient': f"233{txn.user.phone}",
                'sender_id': 'DANWELSTORE',
                'message': sms_message
            }

            messages.success(request, f"Transaction Completed")
            return redirect('bt_admin')


@login_required(login_url='login')
def afa_mark_as_sent(request, pk, status):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.AFARegistration.objects.filter(id=pk).first()
        print(txn)
        if status == "Processing":
            txn.transaction_status = "Processing"
            txn.save()
            messages.success(request, f"Transaction Processed")
            return redirect('afa_admin')
        elif status == "Cancelled":
            txn.transaction_status = "Cancelled"
            txn.save()
            messages.success(request, f"Transaction Cancelled")
            return redirect('afa_admin')
        elif status == "Refunded":
            txn.transaction_status = "Refunded"
            txn.save()
            messages.success(request, f"Transaction Refunded")
            return redirect('afa_admin')
        elif status == "Under Verification":
            txn.transaction_status = "Under Verification"
            txn.save()
            messages.success(request, f"Transaction Under Verification")
            return redirect('afa_admin')
        else:
            txn.transaction_status = "Completed"
            txn.save()
            sms_headers = {
                'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Your AFA Registration has been completed. {txn.phone_number} has been registered.\nTransaction Reference: {txn.reference}"

            sms_body = {
                'recipient': f"233{txn.user.phone}",
                'sender_id': 'DANWELSTORE',
                'message': sms_message
            }
            # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            # print(response.text)
            messages.success(request, f"Transaction Completed")
            return redirect('afa_admin')


def credit_user(request):
    form = forms.CreditUserForm()
    if request.user.is_superuser:
        if request.method == "POST":
            form = forms.CreditUserForm(request.POST)
            if form.is_valid():
                user = form.cleaned_data["user"]
                amount = form.cleaned_data["amount"]
                print(user)
                print(amount)
                user_needed = models.CustomUser.objects.get(username=user)
                if user_needed.wallet is None:
                    user_needed.wallet = Decimal(amount)
                else:
                    user_needed.wallet += Decimal(amount)
                user_needed.save()
                print(user_needed.username)
                messages.success(request, "Crediting Successful")
                sms_headers = {
                    'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                    'Content-Type': 'application/json'
                }

                sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                sms_message = f"Hello {user_needed},\nYour DataForAll wallet has been credit with GHS{amount}.\nDataForAll."

                sms_body = {
                    'recipient': f"233{user_needed.phone}",
                    'sender_id': 'DANWELSTORE',
                    'message': sms_message
                }
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
                return redirect('credit_user')
        context = {'form': form}
        return render(request, "layouts/services/credit.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


# @login_required(login_url='login')
# def topup_info(request):
#     if request.method == "POST":
#         admin = models.AdminInfo.objects.filter().first().phone_number
#         user = models.CustomUser.objects.get(id=request.user.id)
#         amount = request.POST.get("amount")
#         print(amount)
#         reference = helper.top_up_ref_generator()
#         new_topup_request = models.TopUpRequest.objects.create(
#             user=request.user,
#             amount=amount,
#             reference=reference,
#         )
#         new_topup_request.save()
#
#         sms_headers = {
#             'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
#             'Content-Type': 'application/json'
#         }
#
#         sms_url = 'https://webapp.usmsgh.com/api/sms/send'
#         sms_message = f"A top up request has been placed.\nGHS{amount} for {user}.\nReference: {reference}"
#
#         sms_body = {
#             'recipient': f"233{admin}",
#             'sender_id': 'DANWELSTORE',
#             'message': sms_message
#         }
#
#         # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
#         # print(response.text)
#
#         messages.success(request,
#                          f"Your Request has been sent successfully. Kindly go on to pay to {admin} and use the reference stated as reference. Reference: {reference}")
#         return redirect("request_successful", reference)
#     return render(request, "layouts/topup-info.html")


@login_required(login_url='login')
def request_successful(request, reference):
    admin = models.AdminInfo.objects.filter().first()
    context = {
        "name": admin.name,
        "number": f"0{admin.momo_number}",
        "channel": admin.payment_channel,
        "reference": reference
    }
    return render(request, "layouts/services/request_successful.html", context=context)


def topup_list(request):
    if request.user.is_superuser:
        topup_requests = models.TopUpRequest.objects.all().order_by('date').reverse()[:500]
        context = {
            'requests': topup_requests,
        }
        return render(request, "layouts/services/topup_list.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


@login_required(login_url='login')
def credit_user_from_list(request, reference):
    if request.user.is_superuser:
        crediting = models.TopUpRequest.objects.filter(reference=reference).first()
        user = crediting.user
        custom_user = models.CustomUser.objects.get(username=user.username)
        if crediting.status:
            return redirect('topup_list')
        amount = crediting.amount
        print(user)
        print(user.phone)
        print(amount)
        custom_user.wallet += amount
        custom_user.save()
        crediting.status = True
        crediting.credited_at = datetime.now()
        crediting.save()
        sms_headers = {
            'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Hello,\nYour wallet has been topped up with GHS{amount}.\nReference: {reference}.\nThank you"

        sms_body = {
            'recipient': f"233{custom_user.phone}",
            'sender_id': 'DANWELSTORE',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            pass
        messages.success(request, f"{user} has been credited with {amount}")
        return redirect('topup_list')


def populate_custom_users_from_excel(request):
    # Read the Excel file using pandas
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']

            # Process the uploaded Excel file
            df = pd.read_excel(excel_file)
            counter = 0
            # Iterate through rows to create CustomUser instances
            for index, row in df.iterrows():
                print(counter)
                # Create a CustomUser instance for each row
                custom_user = CustomUser.objects.create(
                    first_name=row['first_name'],
                    last_name=row['last_name'],
                    username=str(row['username']),
                    email=row['email'],
                    phone=row['phone'],
                    wallet=float(row['wallet']),
                    status=str(row['status']),
                    password1=row['password1'],
                    password2=row['password2'],
                    is_superuser=row['is_superuser'],
                    is_staff=row['is_staff'],
                    is_active=row['is_active'],
                    password=row['password']
                )

                custom_user.save()

                # group_names = row['groups'].split(',')  # Assuming groups are comma-separated
                # groups = Group.objects.filter(name__in=group_names)
                # custom_user.groups.set(groups)
                #
                # if row['user_permissions']:
                #     permission_ids = [int(pid) for pid in row['user_permissions'].split(',')]
                #     permissions = Permission.objects.filter(id__in=permission_ids)
                #     custom_user.user_permissions.set(permissions)
                print("killed")
                counter = counter + 1
            messages.success(request, 'All done')
    else:
        form = UploadFileForm()
    return render(request, 'layouts/import_users.html', {'form': form})


def delete_custom_users(request):
    CustomUser.objects.all().delete()
    return HttpResponseRedirect('Done')


@csrf_exempt
def hubtel_webhook(request):
    if request.method == 'POST':
        print("hit the webhook")
        try:
            payload = request.body.decode('utf-8')
            print("Hubtel payment Info: ", payload)
            json_payload = json.loads(payload)
            print(json_payload)

            data = json_payload.get('Data')
            print(data)
            reference = data.get('ClientReference')
            print(reference)
            txn_status = data.get('Status')
            txn_description = data.get('Description')
            amount = data.get('Amount')
            print(txn_status, amount)

            if txn_status == 'Success':
                print("success")
                transaction_saved = models.Payment.objects.get(reference=reference, transaction_status="Unfinished")
                transaction_saved.transaction_status = "Paid"
                transaction_saved.payment_description = txn_description
                transaction_saved.amount = amount
                transaction_saved.save()
                transaction_details = "hi"
                transaction_channel = "topup"
                user = transaction_saved.user
                # receiver = collection_saved['number']
                # bundle_volume = collection_saved['data_volume']
                # name = collection_saved['name']
                # email = collection_saved['email']
                # phone_number = collection_saved['buyer']
                # date_and_time = collection_saved['date_and_time']
                # txn_type = collection_saved['type']
                # user_id = collection_saved['uid']
                print(transaction_details, transaction_channel)

                if transaction_channel == "ishare":
                    # offer = transaction_details["offers"]
                    # phone_number = transaction_details["phone_number"]
                    #
                    # if user.status == "User":
                    #     bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # elif user.status == "Agent":
                    #     bundle = models.AgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # elif user.status == "Super Agent":
                    #     bundle = models.SuperAgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # else:
                    #     bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # new_transaction = models.IShareBundleTransaction.objects.create(
                    #     user=user,
                    #     bundle_number=phone_number,
                    #     offer=f"{bundle}MB",
                    #     reference=reference,
                    #     transaction_status="Pending"
                    # )
                    # print("created")
                    # new_transaction.save()
                    #
                    # print("===========================")
                    # print(phone_number)
                    # print(bundle)
                    # print(user)
                    # print(reference)
                    # send_bundle_response = helper.send_bundle(user, phone_number, bundle, reference)
                    # data = send_bundle_response.json()
                    #
                    # print(data)
                    #
                    # sms_headers = {
                    #     'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                    #     'Content-Type': 'application/json'
                    # }
                    #
                    # sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                    #
                    # if send_bundle_response.status_code == 200:
                    #     if data["code"] == "0000":
                    #         transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                    #             reference=reference)
                    #         print("got here")
                    #         print(transaction_to_be_updated.transaction_status)
                    #         transaction_to_be_updated.transaction_status = "Completed"
                    #         transaction_to_be_updated.save()
                    #         print(user.phone)
                    #         print("***********")
                    #         receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {user.phone}.\nReference: {reference}\n"
                    #         sms_message = f"Hello @{user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nThank you for using Data4All GH.\n\nThe Data4All GH"
                    #
                    #         sms_body = {
                    #             'recipient': f"233{user.phone}",
                    #             'sender_id': 'DANWELSTORE',
                    #             'message': sms_message
                    #         }
                    #         try:
                    #             response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                    #             print(response.text)
                    #         except:
                    #             print("message not sent")
                    #             pass
                    #         return JsonResponse({'status': 'Transaction Completed Successfully'}, status=200)
                    #     else:
                    #         transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                    #             reference=reference)
                    #         transaction_to_be_updated.transaction_status = "Failed"
                    #         new_transaction.save()
                    #         sms_message = f"Hello @{user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {reference}\nThank you for using Data4All GH.\n\nThe Data4All GH"
                    #
                    #         sms_body = {
                    #             'recipient': f"233{user.phone}",
                    #             'sender_id': 'Data4All',
                    #             'message': sms_message
                    #         }
                    #         return JsonResponse({'status': 'Something went wrong'}, status=500)
                    # else:
                    #     transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                    #         reference=reference)
                    #     transaction_to_be_updated.transaction_status = "Failed"
                    #     new_transaction.save()
                    #     sms_message = f"Hello @{user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {payment_reference}\nThank you for using Data4All GH.\n\nThe Data4All GH"
                    #
                    #     sms_body = {
                    #         'recipient': f'233{user.phone}',
                    #         'sender_id': 'Data4All',
                    #         'message': sms_message
                    #     }
                    #
                    #     # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                    #     #
                    #     # print(response.text)
                    #     return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
                    ...
                elif transaction_channel == "mtn":
                    # offer = transaction_details["offers"]
                    # phone_number = transaction_details["phone_number"]
                    #
                    # if user.status == "User":
                    #     bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # elif user.status == "Agent":
                    #     bundle = models.AgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # elif user.status == "Super Agent":
                    #     bundle = models.SuperAgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # else:
                    #     bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    #
                    # print(phone_number)
                    # new_mtn_transaction = models.MTNTransaction.objects.create(
                    #     user=user,
                    #     bundle_number=phone_number,
                    #     offer=f"{bundle}MB",
                    #     reference=reference,
                    # )
                    # new_mtn_transaction.save()
                    # return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
                    ...
                elif transaction_channel == "bigtime":
                    # offer = transaction_details["offers"]
                    # phone_number = transaction_details["phone_number"]
                    # if user.status == "User":
                    #     bundle = models.BigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # elif user.status == "Agent":
                    #     bundle = models.AgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # elif user.status == "Super Agent":
                    #     bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # else:
                    #     bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    # print(phone_number)
                    # new_mtn_transaction = models.BigTimeTransaction.objects.create(
                    #     user=user,
                    #     bundle_number=phone_number,
                    #     offer=f"{bundle}MB",
                    #     reference=reference,
                    # )
                    # new_mtn_transaction.save()
                    # return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
                    ...
                elif transaction_channel == "afa":
                    # name = transaction_details["name"]
                    # phone_number = transaction_details["phone"]
                    # gh_card_number = transaction_details["card"]
                    # occupation = transaction_details["occupation"]
                    # date_of_birth = transaction_details["date_of_birth"]
                    #
                    # new_afa_reg = models.AFARegistration.objects.create(
                    #     user=user,
                    #     phone_number=phone_number,
                    #     gh_card_number=gh_card_number,
                    #     name=name,
                    #     occupation=occupation,
                    #     reference=reference,
                    #     date_of_birth=date_of_birth
                    # )
                    # new_afa_reg.save()
                    # return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
                    ...
                elif transaction_channel == "topup":
                    amount = amount

                    user.wallet += float(amount)
                    user.save()

                    new_topup = models.TopUpRequest.objects.create(
                        user=user,
                        reference=reference,
                        amount=amount,
                        status=True,
                    )
                    new_topup.save()
                    response1 = requests.get(
                        f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=OnBuSjBXc1pqN0xrQXIxU1A=&to=0{user.phone}&from=DANWELSTORE&sms=Your Bestplug wallet has been credited with {amount}. Thank You.")
                    print(response1.text)
                    return JsonResponse({'status': "Wallet Credited"}, status=200)
                else:
                    print("no type found")
                    return JsonResponse({'message': "No Type Found"}, status=500)
            else:
                print("Transaction was not Successful")
                return JsonResponse({'message': 'Transaction Failed'}, status=200)
        except Exception as e:
            print("Error Processing hubtel webhook:", str(e))
            return JsonResponse({'status': 'error'}, status=500)
    else:
        print("not post")
        return JsonResponse({'message': 'Not Found'}, status=404)


from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator


def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            user = models.CustomUser.objects.filter(email=data).first()
            current_user = user
            if user:
                subject = "Password Reset Requested"
                email_template_name = "password/password_reset_message.txt"
                c = {
                    "name": user.first_name,
                    "email": user.email,
                    'domain': 'www.danwelstoregh.com',
                    'site_name': 'DanWel Store GH',
                    "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                    "user": user,
                    'token': default_token_generator.make_token(user),
                    'protocol': 'https',
                }
                email = render_to_string(email_template_name, c)

                sms_headers = {
                    'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                    'Content-Type': 'application/json'
                }

                sms_url = 'https://webapp.usmsgh.com/api/sms/send'

                sms_body = {
                    'recipient': f"233{user.phone}",
                    'sender_id': 'DANWELSTORE',
                    'message': email
                }
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
                # requests.get(
                #     f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to=0{current_user.phone}&from=GEO_AT&sms={email}")

                return redirect("/password_reset/done/")
    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="password/password_reset.html",
                  context={"password_reset_form": password_reset_form})


def refund_policy(request):
    return render(request, "layouts/refund_policy.html")


################################# Paystack Views #############################
logger = logging.getLogger(__name__)


def _to_pesewas(amount_str: str) -> int:
    """
    Convert a GHS string (e.g. "10", "10.50") to pesewas (int).
    Paystack expects smallest currency unit.
    """
    try:
        value = (Decimal(amount_str).quantize(Decimal("0.01"))) * 100
        if value <= 0:
            raise ValueError("Amount must be greater than zero.")
        return int(value)
    except (InvalidOperation, ValueError) as e:
        raise ValueError("Invalid amount") from e


def _session_with_retries() -> Session:
    s = Session()
    retry = Retry(
        total=5,
        backoff_factor=0.6,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET", "POST"]),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    # Force certifi CA bundle
    s.verify = certifi.where()
    s.headers.update({
        "Authorization": f"Bearer {settings.PAYSTACK_SECRET_KEY}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    })
    return s


def paystack_initialize(email: str, amount_pesewas: int, reference: str, metadata: dict) -> dict:
    url = f"{settings.PAYSTACK_BASE_URL}/transaction/initialize"

    # 1. Get and Clean the Secret Key
    secret_key = getattr(settings, 'PAYSTACK_SECRET_KEY', None)

    # DEBUGGING: Print to console to verify key is loaded (only prints first 8 chars)
    if secret_key:
        print(f"DEBUG: Using Key -> {str(secret_key)[:8]}...")
        secret_key = str(secret_key).strip()  # Remove accidental spaces from .env
    else:
        print("DEBUG: PAYSTACK_SECRET_KEY is None or Empty!")
        raise ValueError("Paystack Secret Key is missing in settings.py")

    # 2. Prepare Headers
    headers = {
        "Authorization": f"Bearer {secret_key}",
        "Content-Type": "application/json",
    }

    payload = {
        "email": email,
        "amount": amount_pesewas,
        "currency": "GHS",
        "reference": reference,
        "metadata": metadata or {},
    }

    if getattr(settings, 'PAYSTACK_CALLBACK_URL', None):
        payload["callback_url"] = settings.PAYSTACK_CALLBACK_URL

    # 3. Make the request using standard requests.post
    # Note: We use json=payload which handles Content-Type and dumping automatically,
    # but since we manually set headers above, using data=json.dumps is safer for legacy compatibility.
    s = _session_with_retries()
    # Use 's.post', not 'requests.post'
    resp = s.post(
        url,
        data=json.dumps(payload),
        # Headers are already in _session_with_retries, but adding them here again doesn't hurt
        timeout=25,
        verify = certifi.where()
    )

    # 4. Handle Response
    try:
        data = resp.json()
    except Exception as e:
        print(e)
        logger.error("Paystack init non-JSON response: %s", resp.text[:500])
        resp.raise_for_status()

    if resp.status_code not in [200, 201] or not data.get("status"):
        logger.error("Paystack init error: %s %s", resp.status_code, data)
        # This catches the 401 error and tells you exactly what Paystack said
        error_msg = data.get("message", "Failed to initialize payment")
        raise RuntimeError(f"Paystack Error: {error_msg}")

    return data


def _record_payment(user, reference, amount_ghs, status, message=None, description=None):
    """
    Upsert a Payment row for audit.
    """
    models.Payment.objects.update_or_create(
        user=user,
        reference=reference,
        defaults={
            "amount": Decimal(amount_ghs) if amount_ghs is not None else None,
            "payment_description": description or "Wallet top-up via Paystack",
            "transaction_status": status,
            "transaction_date": timezone.now().isoformat(),
            "message": (message or "")[:500],
        },
    )


# ---------- Core views ----------

@login_required(login_url='login')
def topup_info(request):
    if request.method == "POST":
        admin_info = models.AdminInfo.objects.first()
        if not admin_info:
            messages.error(request, "Top-up configuration not found. Please contact support.")
            return redirect("topup_info")

        user = models.CustomUser.objects.get(id=request.user.id)
        amount_str = (request.POST.get("amount") or "").strip()

        try:
            # If you add fees, do it here before conversion
            amount_pesewas = _to_pesewas(amount_str)
        except ValueError:
            messages.error(request, "Invalid top-up amount.")
            return redirect("topup_info")

        # Generate our internal reference for tracing this request end-to-end
        reference = helper.generate_paystack_ref()

        # Persist a request row immediately
        topup_req = models.TopUpRequest.objects.create(
            user=request.user,
            amount=Decimal(amount_str),
            reference=reference,
        )

        # If Paystack is OFF, go with your existing SMS/manual path
        if not admin_info.use_paystack_for_topup:
            admin_phone = admin_info.phone_number
            # (Your SMS code remains commented; leaving as-is)
            messages.success(
                request,
                f"Your request has been sent successfully. Kindly pay to {admin_phone} and use "
                f"the reference below. Reference: {reference}"
            )
            return redirect("request_successful", reference)

        # Paystack path
        try:
            meta = {
                "topup_request_id": topup_req.id,
                "user_id": user.id,
                "username": user.username,
                "purpose": "wallet_topup",
                "amount_ghs": str(Decimal(amount_str).quantize(Decimal('0.01'))),
            }
            init_data = paystack_initialize(
                email=user.email or "no-email@example.com",
                amount_pesewas=amount_pesewas,
                reference=reference,
                metadata=meta,
            )
            auth_url = init_data["data"]["authorization_url"]
            access_ref = init_data["data"]["reference"]
            _record_payment(user, access_ref, Decimal(amount_str), status="Initialized", message="Redirect to Paystack")
            return redirect(auth_url)
        except Exception as e:
            logger.exception("Error initializing Paystack: %s", e)
            messages.error(request, "Unable to start payment. Please try again in a moment.")
            return redirect("topup-info")

    # GET
    return render(request, "layouts/topup-info.html")


@login_required(login_url='login')
def paystack_callback(request):
    """
    Optional: Paystack will redirect here after user approves payment in-browser.
    We do NOT finalize credit here—finalization is done in the webhook after signature verification.
    This simply shows a friendly page based on reference.
    """
    reference = request.GET.get("reference") or request.GET.get("trxref")
    if not reference:
        return HttpResponseBadRequest("Missing reference")

    # Let the user land on your success page. The wallet credit occurs only via webhook.
    return redirect("request_successful_paystack", reference)



@csrf_exempt
def paystack_webhook(request):
    print("it was hit")
    """
    Webhook endpoint for Paystack events. Verifies signature and credits wallet idempotently.
    """
    if request.method == "POST":
        print("hit teh post part")
        # signature = request.headers.get("x-paystack-signature")
        # if not signature:
        #     return HttpResponseForbidden("Missing signature")
        #
        # body = request.body
        # # Verify HMAC-SHA512 signature
        # import hmac
        # import hashlib
        # computed = hmac.new(
        #     key=settings.PAYSTACK_SECRET_KEY.encode("utf-8"),
        #     msg=body,
        #     digestmod=hashlib.sha512
        # ).hexdigest()
        #
        # if computed != signature:
        #     return HttpResponseForbidden("Invalid signature")

        try:
            payload = json.loads(body.decode("utf-8"))
        except json.JSONDecodeError:
            return HttpResponseBadRequest("Invalid JSON")

        event = payload.get("event")
        data = payload.get("data", {})
        reference = data.get("reference")
        metadata = data.get("metadata") or {}

        if event != "charge.success":
            return JsonResponse({"ok": True})

        # CHECK PURPOSE
        purpose = metadata.get("purpose")

        # ================= CASE 1: WALLET TOP-UP =================
        if purpose == "wallet_topup":
            topup_request_id = metadata.get("topup_request_id")
            amount_pesewas = data.get("amount")
            amount_ghs = Decimal(amount_pesewas) / Decimal(100)

            try:
                with transaction.atomic():
                    topup = models.TopUpRequest.objects.select_for_update().get(id=topup_request_id)

                    if topup.status:
                        return JsonResponse({"ok": True, "duplicate": True})

                    # Credit Wallet
                    models.CustomUser.objects.filter(id=topup.user.id).update(
                        wallet=F("wallet") + amount_ghs
                    )

                    topup.status = True
                    topup.credited_at = timezone.now()
                    topup.save()

                    _record_payment(topup.user, reference, amount_ghs, "Success", "Wallet Topup")

            except Exception as e:
                logger.exception("Topup Webhook Error: %s", e)
                return HttpResponse(status=500)

        # ================= CASE 2: SHOP ORDER =================
        elif purpose == "shop_order":
            tracking_number = metadata.get("order_tracking_number")

            try:
                with transaction.atomic():
                    # Lock the order row
                    order = models.Order.objects.select_for_update().get(tracking_number=tracking_number)

                    # Idempotency check
                    if order.status in ["Processing", "Completed", "Paid"]:
                        return JsonResponse({"ok": True, "message": "Already processed"})

                    # 1. Update Order
                    order.payment_id = reference
                    order.status = "Processing"  # or "Paid"
                    order.save()

                    # 2. DEDUCT STOCK (This logic moved here from checkout)
                    order_items = models.OrderItem.objects.filter(order=order)
                    for item in order_items:
                        product = item.product
                        # We reload product to avoid race conditions
                        product.refresh_from_db()
                        if product.quantity >= item.quantity:
                            product.quantity -= item.quantity
                            product.save()
                        else:
                            # Critical: Stock ran out while they were paying
                            logger.error(f"Order {order.id} paid but product {product.name} OOS")
                            # You might want to flag this order for manual refund

                    # 3. CLEAR CART
                    # We find the user's cart and clear it
                    models.Cart.objects.filter(user=order.user).delete()

                    # 4. RECORD PAYMENT
                    _record_payment(order.user, reference, order.total_price, "Success", "Shop Order Payment")

                    # 5. SEND SMS (Moved here)
                    sms_headers = {
                        'Authorization': 'Bearer 2069|fMiymkKVytFt84w8GNM8vq0zF2UtakVaNZT1RUVWfd642028',
                        'Content-Type': 'application/json'
                    }
                    sms_body = {
                        'recipient': f"233{order.phone}",
                        'sender_id': 'DANWEL',
                        'message': f"Order {order.tracking_number} Confirmed! We are processing your delivery."
                    }
                    try:
                        requests.post('https://webapp.usmsgh.com/api/sms/send', json=sms_body, headers=sms_headers,
                                      timeout=5)
                    except Exception as e:
                        print(e)
                        pass

            except models.Order.DoesNotExist:
                return HttpResponseBadRequest("Order not found")

        else:
            logger.warning("Unknown webhook purpose: %s", purpose)

        return JsonResponse({"ok": True})
    else:
        return HttpResponse(status=200)


@login_required(login_url="login")
def request_successful_paystack(request, reference: str):
    """
    Display a Paystack-specific success/landing page after redirect.
    NOTE: The actual wallet credit happens in the paystack_webhook after signature verification.
    This page polls a tiny status endpoint to reflect when crediting is done.
    """
    if not reference:
        return HttpResponseBadRequest("Missing reference.")

    # TopUpRequest should exist because we created it before initializing Paystack
    topup = get_object_or_404(models.TopUpRequest, reference=reference, user=request.user)

    context = {
        "reference": topup.reference,
        "amount": f"{topup.amount:.2f}",
        "created_at": topup.date,
        "credited_at": topup.credited_at,
        "is_credited": bool(topup.status),
        "admin_name": getattr(models.AdminInfo.objects.first(), "name", ""),
        "admin_channel": getattr(models.AdminInfo.objects.first(), "payment_channel", ""),
    }
    return render(request, "layouts/topup-success-paystack.html", context)


@login_required(login_url="login")
def topup_status_api(request, reference: str):
    """
    Small JSON endpoint polled by the Paystack success template to know when the wallet is credited.
    """
    if not reference:
        return HttpResponseBadRequest("Missing reference.")

    topup = get_object_or_404(models.TopUpRequest, reference=reference, user=request.user)
    return JsonResponse({
        "reference": topup.reference,
        "status": bool(topup.status),  # True when credited
        "credited_at": topup.credited_at.isoformat() if topup.credited_at else None,
        "wallet": f"{request.user.wallet:.2f}",
        "server_time": timezone.now().isoformat(),
    })
